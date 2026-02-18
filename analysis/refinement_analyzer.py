"""Phase 5 Refinement Analyzer — multi-metric scoring of TS refinement results.

After TS runs the refinement optimization, this reads the output CSVs and ranks
the top 50 individual parameter combos (across all plateaus) using:
  - MAR Ratio (CAGR / Max Drawdown)
  - Sharpe Ratio approximation
  - Sortino Ratio approximation
  - Profit Factor
  - Profit/Drawdown ratio
  - Robustness (% profitable combinations in source plateau)

Uses rank-sum composite scoring (lower = better), computed per strategy
to prevent a dominant strategy from pushing all others to bottom ranks.

Usage:
    python -m analysis.refinement_analyzer --symbol SMH
    python -m analysis.refinement_analyzer --symbol SMH --top 50
    python -m analysis.refinement_analyzer --symbol SMH --details ATR-D-v1.5-L 1
"""
from __future__ import annotations

import argparse
import glob
import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BACKTEST_YEARS = 10.0
INITIAL_CAPITAL = 10_000.0


def _calculate_cagr(net_profit: float, years: float = BACKTEST_YEARS,
                    initial_capital: float = INITIAL_CAPITAL) -> float:
    ending_value = initial_capital + net_profit
    if ending_value <= 0 or initial_capital <= 0:
        return 0.0
    return ((ending_value / initial_capital) ** (1 / years) - 1) * 100


def _get_parameter_columns(df: pd.DataFrame) -> list[str]:
    metric_prefixes = ["All:", "Long:", "Short:", "TestNum"]
    return [col for col in df.columns
            if not any(col.startswith(p) for p in metric_prefixes)]


def load_refinement_results(symbol_folder: str) -> dict:
    """Load all Refinement_*.csv result files from a symbol folder."""
    files = []
    for pattern_prefix in ["Refinement_"]:
        found = glob.glob(os.path.join(symbol_folder, f"{pattern_prefix}*.csv"))
        files.extend(f for f in found if "rank" in os.path.basename(f).lower()
                     and f not in files)

    print(f"Found {len(files)} refinement result files")

    results = {}
    for filepath in sorted(files):
        filename = os.path.basename(filepath)
        parts = filename.replace("Refinement_", "").replace("Phase5_Refine_", "").replace(".csv", "")

        rank_split = parts.rsplit("_rank", 1)
        if len(rank_split) != 2:
            print(f"  Warning: Could not parse filename: {filename}")
            continue

        strategy_name = rank_split[0]
        rank_part = rank_split[1]
        rank = int(rank_part.split("_")[0]) if "_" in rank_part else int(rank_part)

        try:
            df = pd.read_csv(filepath)
            config_params = _load_config_params(symbol_folder, strategy_name, rank)

            results[(strategy_name, rank)] = {
                "filepath": filepath,
                "strategy": strategy_name,
                "rank": rank,
                "data": df,
                "test_count": len(df),
                "config_params": config_params,
            }
            print(f"  Loaded: {strategy_name} rank {rank} ({len(df)} tests)")
        except Exception as e:
            print(f"  Error loading {filename}: {e}")

    return results


def _load_config_params(symbol_folder: str, strategy_name: str,
                        rank: int) -> dict | None:
    """Load range CSV to get all parameters (T1 vary, T2/T3 fixed)."""
    config_path = os.path.join(symbol_folder,
                               f"phase5_refine_{strategy_name}_rank{rank}.csv")
    if not os.path.exists(config_path):
        return None
    try:
        df = pd.read_csv(config_path)
        params = {}
        for _, row in df.iterrows():
            params[row["ParameterName"]] = {
                "min": row["Min"], "max": row["Max"],
                "is_fixed": row["Min"] == row["Max"],
                "fixed_value": row["Min"] if row["Min"] == row["Max"] else None,
            }
        return params
    except Exception:
        return None


def _compute_combo_metrics(row: pd.Series, param_cols: list[str],
                           strategy_name: str, plateau_rank: int,
                           config_params: dict | None,
                           plateau_robustness: float,
                           plateau_sharpe: float) -> dict:
    """Compute all metrics for a single parameter combo."""
    net_profit = row.get("All: Net Profit", 0)
    max_dd = abs(row.get("All: Max Intraday Drawdown", 0))
    total_trades = row.get("All: Total Trades", 0)
    win_rate = row.get("All: % Profitable", 50)
    avg_losing = abs(row.get("All: Avg Losing Trade", 0))

    analysis = {
        "strategy": strategy_name,
        "plateau_rank": plateau_rank,
        "net_profit": net_profit,
        "max_drawdown": max_dd,
        "profit_factor": row.get("All: ProfitFactor", 0),
        "total_trades": total_trades,
        "win_rate": win_rate,
        "avg_trade": row.get("All: Avg Trade", 0),
        "avg_winning": row.get("All: Avg Winning Trade", 0),
        "avg_losing": avg_losing,
        "robustness": plateau_robustness,
        "plateau_sharpe": plateau_sharpe,
    }

    # CAGR
    cagr = _calculate_cagr(net_profit)
    analysis["cagr_pct"] = cagr

    # MAR Ratio
    peak_equity = INITIAL_CAPITAL + net_profit + max_dd
    if max_dd > 0 and peak_equity > 0:
        max_dd_pct = (max_dd / peak_equity) * 100
        analysis["mar_ratio"] = cagr / max_dd_pct if max_dd_pct > 0 else 0
    else:
        analysis["mar_ratio"] = cagr if cagr > 0 else 0

    # Profit / Drawdown
    analysis["profit_dd_ratio"] = net_profit / max_dd if max_dd > 0 else 0

    # Sharpe approximation
    if total_trades > 0 and max_dd > 0:
        analysis["sharpe_ratio"] = (net_profit / max_dd) * np.sqrt(total_trades / BACKTEST_YEARS)
    else:
        analysis["sharpe_ratio"] = 0

    # Sortino approximation
    if avg_losing > 0 and total_trades > 0:
        downside_dev = avg_losing * np.sqrt(total_trades * (1 - win_rate / 100))
        analysis["sortino_ratio"] = net_profit / downside_dev if downside_dev > 0 else 0
    else:
        analysis["sortino_ratio"] = 0

    # Risk-averse metrics
    avg_winning = row.get("All: Avg Winning Trade", 0)
    avg_trade = row.get("All: Avg Trade", 0)

    # Tail Ratio: avg_winning / |avg_losing| — measures stop discipline
    if avg_losing > 0:
        analysis["tail_ratio"] = avg_winning / avg_losing
    else:
        analysis["tail_ratio"] = avg_winning if avg_winning > 0 else 0

    # Expectancy Ratio: avg_trade / |avg_losing| — profit per unit of risk
    if avg_losing > 0:
        analysis["expectancy_ratio"] = avg_trade / avg_losing
    else:
        analysis["expectancy_ratio"] = avg_trade if avg_trade > 0 else 0

    # Risk-Adjusted Expectancy: avg_trade / trade_std_dev (trade-level Sharpe)
    win_pct = win_rate / 100
    loss_pct = 1 - win_pct
    trade_variance = (win_pct * (avg_winning - avg_trade) ** 2
                      + loss_pct * (-avg_losing - avg_trade) ** 2)
    trade_std_dev = np.sqrt(trade_variance) if trade_variance > 0 else 0
    if trade_std_dev > 0:
        analysis["risk_adj_expectancy"] = avg_trade / trade_std_dev
    else:
        analysis["risk_adj_expectancy"] = 0

    # Build complete parameter set (T1 optimized + T2/T3 fixed)
    combo_params = {col: row[col] for col in param_cols}
    complete_params = {}
    if config_params:
        for param_name, param_info in config_params.items():
            if param_info["is_fixed"]:
                complete_params[param_name] = param_info["fixed_value"]
            elif param_name in combo_params:
                complete_params[param_name] = combo_params[param_name]
            else:
                complete_params[param_name] = (param_info["min"] + param_info["max"]) / 2
    else:
        complete_params = dict(combo_params)

    analysis["all_params"] = complete_params
    return analysis


def get_top_combos(results: dict, top_n: int = 50) -> list[dict]:
    """Get top N combos by net profit across all plateaus, with metrics."""
    all_combos = []

    for key, data in sorted(results.items()):
        strategy_name, rank = key
        df = data["data"]
        config_params = data.get("config_params") or {}
        param_cols = _get_parameter_columns(df)

        if "All: Net Profit" not in df.columns or len(df) == 0:
            continue

        # Plateau-level stats for robustness context
        profitable = df[df["All: Net Profit"] > 0]
        robustness = len(profitable) / len(df) * 100

        avg_profit = df["All: Net Profit"].mean()
        std_profit = df["All: Net Profit"].std()
        plateau_sharpe = avg_profit / std_profit if std_profit > 0 else 0

        # Get top combos from this plateau (take more than needed, dedup later)
        top_rows = df.nlargest(top_n, "All: Net Profit")

        for _, row in top_rows.iterrows():
            combo = _compute_combo_metrics(
                row, param_cols, strategy_name, rank,
                config_params, robustness, plateau_sharpe,
            )
            all_combos.append(combo)

    # Sort by net profit and take top N
    all_combos.sort(key=lambda x: x["net_profit"], reverse=True)
    return all_combos[:top_n]


def get_top_combos_by_strategy(results: dict,
                               top_n: int = 50) -> dict[str, list[dict]]:
    """Get top N combos per strategy, keyed by strategy name.

    Unlike get_top_combos() which pools everything, this ensures each strategy
    gets its own top-N ranking so a dominant strategy can't crowd others out.
    """
    # Group results by strategy
    by_strategy: dict[str, list[dict]] = {}

    for key, data in sorted(results.items()):
        strategy_name, rank = key
        df = data["data"]
        config_params = data.get("config_params") or {}
        param_cols = _get_parameter_columns(df)

        if "All: Net Profit" not in df.columns or len(df) == 0:
            continue

        profitable = df[df["All: Net Profit"] > 0]
        robustness = len(profitable) / len(df) * 100

        avg_profit = df["All: Net Profit"].mean()
        std_profit = df["All: Net Profit"].std()
        plateau_sharpe = avg_profit / std_profit if std_profit > 0 else 0

        top_rows = df.nlargest(top_n, "All: Net Profit")

        for _, row in top_rows.iterrows():
            combo = _compute_combo_metrics(
                row, param_cols, strategy_name, rank,
                config_params, robustness, plateau_sharpe,
            )
            by_strategy.setdefault(strategy_name, []).append(combo)

    # Sort and trim each strategy's combos
    for strategy_name in by_strategy:
        combos = by_strategy[strategy_name]
        combos.sort(key=lambda x: x["net_profit"], reverse=True)
        by_strategy[strategy_name] = combos[:top_n]

    return by_strategy


def calculate_rank_sum(combos: list[dict]) -> list[dict]:
    """Calculate multi-factor rank-sum score. Lower = better."""
    metrics = [
        ("net_profit", False),
        ("mar_ratio", False),
        ("profit_dd_ratio", False),
        ("sharpe_ratio", False),
        ("profit_factor", False),
        ("robustness", False),
    ]

    for c in combos:
        c["ranks"] = {}
        c["rank_sum"] = 0

    for metric, lower_is_better in metrics:
        sorted_combos = sorted(
            combos,
            key=lambda x: x.get(metric, 0) if not lower_is_better else -x.get(metric, 0),
            reverse=True,
        )
        for rank, c in enumerate(sorted_combos, 1):
            c["ranks"][metric] = rank
            c["rank_sum"] += rank

    return combos


def calculate_risk_adjusted_rank_sum(combos: list[dict]) -> list[dict]:
    """Calculate risk-averse rank-sum score. Lower = better.

    Drops net_profit and profit_dd_ratio (both reward absolute return size).
    Adds sortino_ratio, tail_ratio, risk_adj_expectancy (reward risk discipline).
    """
    metrics = [
        ("sortino_ratio", False),
        ("mar_ratio", False),
        ("tail_ratio", False),
        ("risk_adj_expectancy", False),
        ("profit_factor", False),
        ("robustness", False),
    ]

    for c in combos:
        c.setdefault("risk_adj_ranks", {})
        c["risk_adj_rank_sum"] = 0

    for metric, lower_is_better in metrics:
        sorted_combos = sorted(
            combos,
            key=lambda x: x.get(metric, 0) if not lower_is_better else -x.get(metric, 0),
            reverse=True,
        )
        for rank, c in enumerate(sorted_combos, 1):
            c["risk_adj_ranks"][metric] = rank
            c["risk_adj_rank_sum"] += rank

    return combos


def generate_report(results: dict, symbol_folder: str,
                    top_n: int = 50) -> dict[str, list[dict]]:
    """Get top combos per strategy, rank within each, and save results.

    Returns dict[strategy_name, list[dict]] with rank_sum computed per strategy.
    """
    by_strategy = get_top_combos_by_strategy(results, top_n)

    # Rank within each strategy independently
    for strategy_name, combos in by_strategy.items():
        calculate_rank_sum(combos)
        calculate_risk_adjusted_rank_sum(combos)
        combos.sort(key=lambda x: x["rank_sum"])

    # Print ranked summary for each strategy
    for strategy_name, combos in sorted(by_strategy.items()):
        print(f"\n{'='*110}")
        print(f"PHASE 5 REFINEMENT ANALYSIS — {strategy_name} "
              f"— Top {len(combos)} Combos")
        print(f"{'='*110}")
        header = (f"{'#':<4} {'Strategy':<22} {'Plat':>4} {'Net Profit':>12} "
                  f"{'MAR':>6} {'P/DD':>6} {'Sharpe':>7} {'PF':>8} {'Rob%':>5} "
                  f"{'Trades':>6} {'Win%':>5} {'RankSum':>8}")
        print(header)
        print("-" * 110)

        for i, c in enumerate(combos, 1):
            print(f"{i:<4} {c['strategy']:<22} {c['plateau_rank']:>4} "
                  f"${c.get('net_profit', 0):>11,.0f} "
                  f"{c.get('mar_ratio', 0):>6.2f} "
                  f"{c.get('profit_dd_ratio', 0):>6.2f} "
                  f"{c.get('sharpe_ratio', 0):>7.2f} "
                  f"{c.get('profit_factor', 0):>8.2f} "
                  f"{c.get('robustness', 0):>5.1f} "
                  f"{c.get('total_trades', 0):>6.0f} "
                  f"{c.get('win_rate', 0):>5.1f} "
                  f"{c.get('rank_sum', 0):>8}")

    # Save outputs
    all_combos = []
    for combos in by_strategy.values():
        all_combos.extend(combos)
    _save_metrics_csv(all_combos, symbol_folder)
    _save_best_parameters(all_combos, symbol_folder)

    return by_strategy


def _save_metrics_csv(combos: list[dict], symbol_folder: str):
    """Save all scoring metrics to CSV."""
    rows = []
    for c in combos:
        row = {
            "Strategy": c["strategy"],
            "Plateau_Rank": c["plateau_rank"],
            "Rank_Sum": c.get("rank_sum", 0),
            "Risk_Adj_Rank_Sum": c.get("risk_adj_rank_sum", 0),
            "Net_Profit": c.get("net_profit", 0),
            "Max_Drawdown": c.get("max_drawdown", 0),
            "CAGR_Pct": c.get("cagr_pct", 0),
            "Total_Trades": c.get("total_trades", 0),
            "Win_Rate": c.get("win_rate", 0),
            "MAR_Ratio": c.get("mar_ratio", 0),
            "Profit_DD_Ratio": c.get("profit_dd_ratio", 0),
            "Sharpe_Ratio": c.get("sharpe_ratio", 0),
            "Sortino_Ratio": c.get("sortino_ratio", 0),
            "Profit_Factor": c.get("profit_factor", 0),
            "Tail_Ratio": c.get("tail_ratio", 0),
            "Expectancy_Ratio": c.get("expectancy_ratio", 0),
            "Risk_Adj_Expectancy": c.get("risk_adj_expectancy", 0),
            "Robustness_Pct": c.get("robustness", 0),
            "Rank_NetProfit": c.get("ranks", {}).get("net_profit", 0),
            "Rank_MAR": c.get("ranks", {}).get("mar_ratio", 0),
            "Rank_ProfitDD": c.get("ranks", {}).get("profit_dd_ratio", 0),
            "Rank_Sharpe": c.get("ranks", {}).get("sharpe_ratio", 0),
            "Rank_PF": c.get("ranks", {}).get("profit_factor", 0),
            "Rank_Robustness": c.get("ranks", {}).get("robustness", 0),
            "Plateau_Sharpe": c.get("plateau_sharpe", 0),
        }
        for param, value in c.get("all_params", {}).items():
            row[f"Param_{param}"] = value
        rows.append(row)

    df = pd.DataFrame(rows).sort_values("Rank_Sum")
    output_path = os.path.join(symbol_folder, "phase5_scoring_metrics.csv")
    df.to_csv(output_path, index=False)
    print(f"Scoring metrics saved to: {output_path}")


def _save_best_parameters(combos: list[dict], symbol_folder: str):
    """Save the #1 ranked combo per strategy (both return-focused and risk-averse)."""
    strategy_best = {}
    strategy_risk_best = {}
    for c in combos:
        s = c["strategy"]
        if s not in strategy_best or c["rank_sum"] < strategy_best[s]["rank_sum"]:
            strategy_best[s] = c
        ra = c.get("risk_adj_rank_sum", float("inf"))
        if s not in strategy_risk_best or ra < strategy_risk_best[s].get("risk_adj_rank_sum", float("inf")):
            strategy_risk_best[s] = c

    rows = []
    for strategy in sorted(strategy_best.keys()):
        c = strategy_best[strategy]
        row = {
            "Strategy": strategy,
            "Type": "Return-Focused",
            "Best_Plateau_Rank": c["plateau_rank"],
            "Rank_Sum": c.get("rank_sum", 0),
            "Risk_Adj_Rank_Sum": c.get("risk_adj_rank_sum", 0),
            "Net_Profit": c.get("net_profit", 0),
            "MAR_Ratio": c.get("mar_ratio", 0),
            "Profit_DD_Ratio": c.get("profit_dd_ratio", 0),
            "Sharpe_Ratio": c.get("sharpe_ratio", 0),
            "Profit_Factor": c.get("profit_factor", 0),
            "Tail_Ratio": c.get("tail_ratio", 0),
            "Expectancy_Ratio": c.get("expectancy_ratio", 0),
            "Risk_Adj_Expectancy": c.get("risk_adj_expectancy", 0),
            "Robustness_Pct": c.get("robustness", 0),
            "Max_Drawdown": c.get("max_drawdown", 0),
            "Total_Trades": c.get("total_trades", 0),
            "Win_Rate": c.get("win_rate", 0),
        }
        for param, value in c.get("all_params", {}).items():
            row[param] = value
        rows.append(row)

        # Add risk-averse best if different from return-focused best
        rc = strategy_risk_best[strategy]
        if rc is not c:
            rrow = {
                "Strategy": strategy,
                "Type": "Risk-Averse",
                "Best_Plateau_Rank": rc["plateau_rank"],
                "Rank_Sum": rc.get("rank_sum", 0),
                "Risk_Adj_Rank_Sum": rc.get("risk_adj_rank_sum", 0),
                "Net_Profit": rc.get("net_profit", 0),
                "MAR_Ratio": rc.get("mar_ratio", 0),
                "Profit_DD_Ratio": rc.get("profit_dd_ratio", 0),
                "Sharpe_Ratio": rc.get("sharpe_ratio", 0),
                "Profit_Factor": rc.get("profit_factor", 0),
                "Tail_Ratio": rc.get("tail_ratio", 0),
                "Expectancy_Ratio": rc.get("expectancy_ratio", 0),
                "Risk_Adj_Expectancy": rc.get("risk_adj_expectancy", 0),
                "Robustness_Pct": rc.get("robustness", 0),
                "Max_Drawdown": rc.get("max_drawdown", 0),
                "Total_Trades": rc.get("total_trades", 0),
                "Win_Rate": rc.get("win_rate", 0),
            }
            for param, value in rc.get("all_params", {}).items():
                rrow[param] = value
            rows.append(rrow)

    df = pd.DataFrame(rows).sort_values(["Strategy", "Type"])
    output_path = os.path.join(symbol_folder, "phase5_best_parameters.csv")
    df.to_csv(output_path, index=False)
    print(f"Best parameters saved to: {output_path}")


# ---------------------------------------------------------------------------
# Excel styling constants
# ---------------------------------------------------------------------------
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9",
                           fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00",
                           fill_type="solid")
_LIGHT_BLUE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6",
                               fill_type="solid")
_LIGHT_GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90",
                                fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_IMPORTANT_COLS = {
    "Net_Profit", "Max_Drawdown", "CAGR_Pct", "MAR_Ratio",
    "Profit_DD_Ratio", "Sharpe_Ratio", "Profit_Factor",
    "Robustness_Pct", "Rank_Sum", "Risk_Adj_Rank_Sum",
}
_CURRENCY_COLS = {"Net_Profit", "Max_Drawdown", "Avg_Trade"}
_DECIMAL_COLS = {
    "CAGR_Pct", "MAR_Ratio", "Profit_DD_Ratio", "Sharpe_Ratio",
    "Sortino_Ratio", "Plateau_Sharpe", "Profit_Factor", "Win_Rate",
    "Robustness_Pct", "Tail_Ratio", "Expectancy_Ratio",
    "Risk_Adj_Expectancy",
}


def _combos_to_dataframe(combos: list[dict]) -> pd.DataFrame:
    """Convert combo dicts to a DataFrame for Excel writing."""
    rows = []
    best_rank_sum = min(c["rank_sum"] for c in combos) if combos else 0
    best_risk_adj = min(c.get("risk_adj_rank_sum", float("inf"))
                        for c in combos) if combos else 0

    for c in combos:
        is_best = c["rank_sum"] == best_rank_sum
        is_risk_best = c.get("risk_adj_rank_sum", float("inf")) == best_risk_adj
        row = {
            "Strategy": c["strategy"],
            "Plateau_Rank": c["plateau_rank"],
            "Is_Best": is_best,
            "Is_Risk_Best": is_risk_best and not is_best,
            "Rank_Sum": c.get("rank_sum", 0),
            "Risk_Adj_Rank_Sum": c.get("risk_adj_rank_sum", 0),
            "Net_Profit": c.get("net_profit", 0),
            "Max_Drawdown": c.get("max_drawdown", 0),
            "CAGR_Pct": c.get("cagr_pct", 0),
            "MAR_Ratio": c.get("mar_ratio", 0),
            "Profit_DD_Ratio": c.get("profit_dd_ratio", 0),
            "Sharpe_Ratio": c.get("sharpe_ratio", 0),
            "Sortino_Ratio": c.get("sortino_ratio", 0),
            "Plateau_Sharpe": c.get("plateau_sharpe", 0),
            "Profit_Factor": c.get("profit_factor", 0),
            "Tail_Ratio": c.get("tail_ratio", 0),
            "Expectancy_Ratio": c.get("expectancy_ratio", 0),
            "Risk_Adj_Expectancy": c.get("risk_adj_expectancy", 0),
            "Win_Rate": c.get("win_rate", 0),
            "Robustness_Pct": c.get("robustness", 0),
            "Total_Trades": c.get("total_trades", 0),
            "Avg_Trade": c.get("avg_trade", 0),
            "Rank_NetProfit": c.get("ranks", {}).get("net_profit", 0),
            "Rank_MAR": c.get("ranks", {}).get("mar_ratio", 0),
            "Rank_ProfitDD": c.get("ranks", {}).get("profit_dd_ratio", 0),
            "Rank_Sharpe": c.get("ranks", {}).get("sharpe_ratio", 0),
            "Rank_PF": c.get("ranks", {}).get("profit_factor", 0),
            "Rank_Robustness": c.get("ranks", {}).get("robustness", 0),
        }
        for param, value in c.get("all_params", {}).items():
            row[f"Param_{param}"] = value
        rows.append(row)

    return pd.DataFrame(rows).sort_values("Rank_Sum")


def _write_sheet(ws, df: pd.DataFrame):
    """Write a DataFrame to an openpyxl worksheet with standard formatting."""
    columns = [c for c in df.columns if c not in ("Is_Best", "Is_Risk_Best")]

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        if col_name in _IMPORTANT_COLS:
            cell.fill = _YELLOW_FILL
        else:
            cell.fill = _HEADER_FILL

    # Data rows
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        is_best_row = row_data.get("Is_Best", False)
        is_risk_best_row = row_data.get("Is_Risk_Best", False)

        for col_idx, col_name in enumerate(columns, 1):
            value = row_data[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER

            if pd.isna(value) or (isinstance(value, float) and np.isinf(value)):
                if col_name == "Profit_Factor":
                    cell.value = "\u221E"
                else:
                    cell.value = 0
            else:
                cell.value = value

            if col_name in _CURRENCY_COLS:
                cell.number_format = '"$"#,##0'
            elif col_name in _DECIMAL_COLS:
                cell.number_format = "0.00"
            elif col_name.startswith("Param_") and isinstance(value, float):
                cell.number_format = "0.00"

            if is_best_row:
                cell.fill = _LIGHT_BLUE_FILL
            elif is_risk_best_row:
                cell.fill = _LIGHT_GREEN_FILL
            elif col_name in _IMPORTANT_COLS:
                cell.fill = _YELLOW_FILL

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        for row_idx in range(2, len(df) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max_length + 2, 20)

    ws.freeze_panes = "A2"


def save_combined_excel(by_strategy: dict[str, list[dict]],
                        symbol_folder: str, symbol: str):
    """Save combined analysis to Excel with Overview + per-strategy sheets."""
    wb = Workbook()

    # --- Overview sheet: best 20 combos from each strategy ---
    ws_overview = wb.active
    ws_overview.title = "Overview"

    overview_combos = []
    for strategy_name in sorted(by_strategy.keys()):
        combos = by_strategy[strategy_name]
        overview_combos.extend(combos[:20])

    if overview_combos:
        overview_df = _combos_to_dataframe(overview_combos)
        _write_sheet(ws_overview, overview_df)

    # --- Per-strategy sheets ---
    for strategy_name in sorted(by_strategy.keys()):
        combos = by_strategy[strategy_name]
        if not combos:
            continue

        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = strategy_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        df = _combos_to_dataframe(combos)
        _write_sheet(ws, df)

    output_path = os.path.join(symbol_folder, f"phase5_combined_{symbol}.xlsx")
    wb.save(output_path)
    print(f"Combined Excel saved to: {output_path}")


def show_plateau_details(results: dict, strategy_name: str, rank: int):
    """Show detailed analysis for a specific plateau."""
    key = (strategy_name, rank)
    if key not in results:
        print(f"Plateau not found: {strategy_name} rank {rank}")
        return

    data = results[key]
    df = data["data"]
    param_cols = _get_parameter_columns(df)

    print(f"{'='*80}")
    print(f"DETAILED ANALYSIS: {strategy_name} Rank {rank}")
    print(f"{'='*80}")
    print(f"Total tests: {len(df)}")
    print(f"Parameters: {', '.join(param_cols)}")

    if "All: Net Profit" in df.columns:
        print("Top 10 by Net Profit:")
        top10 = df.nlargest(10, "All: Net Profit")
        display_cols = param_cols + ["All: Net Profit", "All: Return on Account", "All: ProfitFactor"]
        display_cols = [c for c in display_cols if c in df.columns]
        print(top10[display_cols].to_string(index=False))

    if "All: Net Profit" in df.columns and len(df) > 10:
        top_pct = df.nlargest(max(1, len(df) // 10), "All: Net Profit")
        print("Parameter ranges in top 10%:")
        for col in param_cols:
            print(f"  {col}: {top_pct[col].min():.4f} - {top_pct[col].max():.4f} "
                  f"(best: {top10.iloc[0][col]:.4f})")


def main():
    parser = argparse.ArgumentParser(description="Analyze Phase 5 refinement results")
    parser.add_argument("--symbol", required=True, help="Symbol folder name (e.g., SMH)")
    parser.add_argument("--base-folder", default=r"C:\Projects\numba-wfo\results",
                        help="Base folder path")
    parser.add_argument("--top", type=int, default=50, help="Number of top combos to include")
    parser.add_argument("--details", nargs=2, metavar=("STRATEGY", "RANK"),
                        help="Show details for specific plateau")

    args = parser.parse_args()

    symbol_folder = os.path.join(args.base_folder, args.symbol)

    if not os.path.exists(symbol_folder):
        print(f"Error: Symbol folder not found: {symbol_folder}")
        return

    print(f"Analyzing refinement results in: {symbol_folder}")

    results = load_refinement_results(symbol_folder)

    if not results:
        print("No refinement results found!")
        return

    if args.details:
        show_plateau_details(results, args.details[0], int(args.details[1]))
        return

    by_strategy = generate_report(results, symbol_folder, args.top)
    save_combined_excel(by_strategy, symbol_folder, args.symbol)

    total_combos = sum(len(combos) for combos in by_strategy.values())
    print(f"\nAnalysis complete. {total_combos} combos ranked "
          f"across {len(by_strategy)} strategies.")
    print("Scoring: MAR>0.5 good, Sharpe>1.0 good, PF>1.5 good")
    print("Rank Sum = sum of ranks across 6 metrics (LOWER = BETTER)")
    print("NOTE: Rank sums are computed WITHIN each strategy (not cross-strategy)")


if __name__ == "__main__":
    main()
